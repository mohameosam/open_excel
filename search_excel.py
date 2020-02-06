#!/usr/bin/python

# Copyright: (c) 2020, Mohamed Abouahmed <mabouahmed@kovarus.com> / Kovarus, mohamedabou99@yahoo.com
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

#!! Uncomment this if you distribute to Ansible Community
#ANSIBLE_METADATA = {'metadata_version': '1.1',
#                    'status': ['preview'],
#                    'supported_by': 'community'}


DOCUMENTATION = '''
---
module: search_excel

short_description: Search for a specific token within an Excel file  (v0.1)

description:
    - Search for a specified string in a specified Excel file and return all matches as a cell-reference in an Ansible list of dictionaries {row_no, col_no}.

author: Mohamed Abouahmed (@mohamedosam)

options:
    src:
        description:
            - The finename of the Excel spreadsheet
        required: true
        type: str

    search_token:
        description:
            - The search token to look for inside the excel sheet
        required: true
        type: str

    search_range:
        description:
            - The start and end cell references to specificy the range at which the search-token will be looked at.
        type: dict
        default: if missing, the module will look in the entire excel sheet (or the entire workbook if the sheet name not specified)
        elements:
            - start_row: Along with "start_col" specifies the cell reference from which the search starts. If omitted, the search will start from the first row in the worksheet
            - start_col: Along with "start_row" specifies the cell reference from which the search starts. If omitted, the search will start from the first column in the worksheet
            - end_row: Along with "end_col" specifies the cell reference at which the search ends. If omitted, the search will end at final row in the worksheet
            - end_col: Along with "end_row" specifies the cell reference at which the search ends. If omitted, the search will end at final column in the worksheet
        
    search_options:
        description:
            - A string carrying all search options
            - The options are as follows:
                - i: ignore case
                - w: whole words
                - x: exact match of the cell content
            - If both "ix" options are specified, the exact non-case sensitive cell search will be processed
            - If both "wx" options are specified, the "w" option will be ignored, and module will search for exact cell content
            - Any other character specified in the search_options will be ignored.
        type: str
        default: if omitted, a case-sensitive search for any occarance of "search_token" will be processed

    sheet_name:
        description:
            - The sheet name to be searched
        type: str
        default: if omitted, all of the workbook sheets are searched
    
requirements:
    - openpyxl Python library must be installed on the Ansible host.
      To install openpyxl, use pip (or pip3) from your linux shell as follows:
      sudo pip install openpyxl  

'''


RETURN = r'''
list:
    description:
        - A list of dictionaries containing cell references where the specified search_token was found. Will return an empty list if search_token was not found
        - Will return empty list if search_token is not found.
    returned: on success
    type: list
    sample:
        - [{row_no:11, col_no:5},{row_no:19, col_no:2}, {row_no:23, col_no:7}]
        - [{row_no:2, col_no:3}]
    elements:
        - row_no: Along with "col_no" specifies the cell reference where the specified search_token was found.
        - col_no: Along with "row_no" specifies the cell reference where the specified search_token was found.    
'''

EXAMPLES = '''

  - name: Read Excel File
    hosts: localhost
    connection: local
    gather_facts: no

    vars:
        search_range:
          start_col: 2 # Search only in the 2nd column of the sheet
          end_col: 2
          
    tasks:
      - name: Looking-up Sandy Cole Info
        search_excel:
          src: "SampleBook.xlsx"
          search_token: "Sandy Cole"
          search_range: "{{ search_range }}"
          sheet_name: "Employee"
          search_options: "ix"
        register: cell_list

      - debug: var=cell_list
      
      # Assuming unique name "Sandy Cole":
      - set_fact:
          read_range: "{{ read_range|default({}) | combine( {'start_row': item['row_no'], 'start_col':1, 'end_row': item['row_no'], 'end_col':4 } ) }}"
        with_items: "{{ cell_list.list }}"      
        when: cell_list.list[0]['row_no'] is defined      
        
      - name: Show Sandy Cole Information
        register: result
        open_excel:
          src: "SampleBook.xlsx"
          op: "r"
          index_by_name: True
          read_range: "{{ read_range }}"
          sheet_name: "Employee"
        when: cell_list.list[0]['row_no'] is defined
        
      - debug: var=result
        when: cell_list.list[0]['row_no'] is defined        

      - debug: msg="Specified search_token was NOT found"
        when: cell_list.list[0]['row_no'] is undefined
        
'''


import openpyxl
import re


#########################################################
###     SEARCH CONTENT: search_xl_content
#########################################################
def search_xl_content(excel_file, search_token, search_range, search_options, sheet_name):
    
    retval = {'list': []}
    
    search_option_i = "i" in search_options
    search_option_w = "w" in search_options
    search_option_x = "x" in search_options
    
    if search_option_i:
        search_token = str(search_token).lower()
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except IOError:
        return (1, "Error accessing excel file [%s]" % excel_file)

    ### Validate & Initialize Parameters
    try:
        start_row = search_range['start_row']
    except Exception as e:
        start_row = 1

    try:
        start_col = search_range['start_col']
    except Exception as e:
        start_col = 1

    try:
        end_row = int(search_range['end_row']) + 1
    except Exception as e:
        end_row = 0

    try:
        end_col = int(search_range['end_col']) + 1
    except Exception as e:
        end_col = 0


    try:        
        sheet_names_list = [sheet_name]
        if not sheet_name:
            sheet_names_list = wb.get_sheet_names()
        
        for asheet_name in sheet_names_list:
            current_sheet = wb.get_sheet_by_name(asheet_name)

            if end_row == 0:
                end_row = current_sheet.max_row + 1

            if end_col == 0:
                end_col = current_sheet.max_column + 1                                        

            for row in range (start_row, end_row):
                for col in range(start_col, end_col):
                    cell_val = str(current_sheet.cell(row=row, column=col).value)
                    if search_option_i:
                        cell_val = cell_val.lower()

                    found_token = False
                    if search_option_x:
                        found_token = search_token == cell_val
                    elif search_option_w:
                        found_token = re.search(r'\b'+ re.escape(search_token) + r'\b', cell_val)
                    else:
                        found_token = search_token in cell_val                        

                    if found_token:
                        temp_dict = {'row_no': row, 'col_no': col}
                        retval['list'].append(dict(temp_dict))
                        
    except Exception as e:
        return(1, "Invalid search parameters specified: %s" % str(e))
    
    return (0, retval)



##################################################################################################################
##################################################################################################################
####     M A I N 
##################################################################################################################
##################################################################################################################

def main():
    module = AnsibleModule(argument_spec = dict(
             src = dict(required=True),
             search_token = dict(type='str',required=True),
             search_range = dict(type='dict',required=False),
             search_options = dict(type='str',default="",required=False),
             sheet_name = dict(required=False)
             ),
             add_file_common_args=True)
    
    ret_code, response = search_xl_content(module.params["src"], module.params["search_token"], module.params["search_range"], module.params["search_options"], module.params["sheet_name"])    
    
    if ret_code:
        module.fail_json(msg=response)
    else:
        module.exit_json(changed=True, **response)

    return ret_code


from ansible.module_utils.basic import *
main()
#
