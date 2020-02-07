#!/usr/bin/python

# Copyright: (c) 2020, Mohamed Abouahmed <mabouahmed@kovarus.com> / Kovarus, mohamedabou99@yahoo.com
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

#!!! Uncomment this if you distribute to Ansible Community
#ANSIBLE_METADATA = {'metadata_version': '1.1',
#                    'status': ['preview'],
#                    'supported_by': 'community'}


DOCUMENTATION = '''
---
module: open_excel.py

short_description: Read an Excel file and register its content into an ansible list of dict (v0.1)

description:
    - Read a specified Excel file and register the Excel file content into an Ansible list of dictionaries. You can specify whether to use
      column headers as the keys, or just use index headers. This is useful when you have non-standard header titles (with invalid reference characters, etc).

author: Mohamed Abouahmed (@mohamedosam)

options:
    src:
        description:
            - The name of the Excel spreadsheet
        required: true
        type: str
    
    dest: 
        description:
            - Only relevant when Excel file is opened for update. The source filename will not be overwritten, instead, the sheet after the update will be saved as this specified destination filename
        type: str
        default: if missing, the generated file will be "<source filename>_updated.xlsx". If the file is opened with "r" mode (read), the "dest" value will be ignored
    
    op:
        description:
            - The operational mode of the file.
            - Valid "op" values:
                - r: to open the file for "read only"
                - w: to write the updates_matrix values in the specified calls. Any old cell value will be overwritten. If cell_row is 0, the module will assume an "a" operational mode
                - a: to append the sheet with a new row at the end of the sheet. The row_no value specified in updates_matrix will be ignored
                - i: insert the updates_matrix values in the row above the row number specified in the row_no parameter. Make sure the row_no values specified in updates_matrix are all the same for valid insert operation. Different row_no values may have unintended results
            - The original file will not be overwritten, instead, another file with a name specified in the paramter "dest" will be created.
        required: true
        type: str
        
    index_by_name:
        description:
            - If true, returns the result values dictionary with keys specified in the excel sheet column headers. When False, keys will carry "col_<n>" format, where <n> is the column number.
            - This option is useful when headers carry titles with invalid dictionary format (have special characters, spaces, etc.)
        type: bool
        default: true
    
    read_range:
        description:
            - The start and end cell references to specificy the range to read from the sheet(s).
        type: dict
        default: if missing, the module will read the entire excel sheet (or the entire workbook if the sheet name not specified)
        elements:
            - start_row: Along with "start_col" specifies the cell reference from which the module will start reading the sheet. If omitted, the read process will start from the first row in the worksheet
            - start_col: Along with "start_row" specifies the cell reference from which the module will start reading the sheet. If omitted, the read process will start from the first column in the worksheet
            - end_row: Along with "end_col" specifies the cell reference at which the module ends reading the sheet content. If omitted, the read process will end at final row in the worksheet
            - end_col: Along with "end_row" specifies the cell reference at which the module ends reading the sheet content. If omitted, the read process will end at final column in the worksheet.

    updates_matrix: A list of dictionary that contains the following:
        cell_row: Along with "cell_col" specifies the cell reference at which "cell_value" will be written
        cell_col: Along with "cell_row" specifies the cell reference at which "cell_value" will be written
        cell_value: The value to write to the cell reference (cell_row, cell_col). 
        
    cell_style: 
        description:
            - The foreground and background color of the cell(s) being updated
            - If omitted, the cell will be updated with the current excisting colors, no change will happen.
        type: dict
        elements:
            - fontColor: A string representing the RGB value of the font color of the cell(s) being updated.
            - bgColor: A string representing the RGB value of the background color of the cell(s) being updated.
            - bold: A boolean value to set the font strike as bold or normal
            - italic: A boolean value to set the font italics True or False
            - underline: A boolean value to set the font underline True or False
            
    sheet_name:
        description:
            - The sheet name to read or updated. This parameter can only be omitted when the Excel sheet is opened as "r" (read-only)
        type: str
        default: sheet_name can only be omitted when the Excel sheet is opened as "r" (read-only). When omitted, the entire worksheet is opened and returned


requirements:
    - openpyxl Python library must be installed on the Ansible host.
      To install openpyxl, use pip (or pip3) from your linux shell as follows:
      sudo pip install openpyxl  

'''


RETURN = r'''
list:
    description: If open for "read-only", the modules returns a list of dictinaries containing the cell value. Will return empty dict if open mode is otherwise
    returned: on success
    type: list
    sample:
    
        "sheet_index_0": [
            {
                "department": "Accounting",
                "employee_name": "John Smith",
                "year_joined": "1999"
            },
            {
                "department": "Engineering",
                "employee_name": "Mo Abou",
                "year_joined": "2008"
            },
            {
                "department": "HR",
                "employee_name": "Sandy Cole",
                "year_joined": "2015"
            },
            {
                "department": "Engineering",
                "employee_name": "Olivia ",
                "year_joined": "2003"
            }
        ],
        "sheet_index_1": [
            {
                "city_code": "408",
                "city_name": "San Jose"
            },
            {
                "city_code": "925",
                "city_name": "Dublin"
            },
            {
                "city_code": "925",
                "city_name": "San Ramon"
            },
            {
                "city_code": "916",
                "city_name": "Sacramento"
            }
        ]
    
    elements:
        - sheet_index_<n>: Reference to the index of the sheet found in the Excel workbook, where <n> refers to the sheet index number (starting from 0)
            - <column header> or col_<n>: If the option "index_by_name" is True, the dictionary key will be the header of the column in the excel sheet. If "index_by_name" is False, the sheet column key will be "col<n>" where <n> is the column number.    

'''


EXAMPLES = '''

vars:
  read_range:
    start_row: 2
    end_row: 40
    start_col: 3
    end_col: 3

  cell_style:
    fontColor: '006100'
    bgColor: 'C6EFCE'
    bold: True
    underline: True

tasks:

  - set_fact:
      updates_matrix: "[{'cell_row': 4, 'cell_col':1, 'cell_value':'D333444' }, {'cell_row': 4, 'cell_col':2, 'cell_value':'John Smith' }, {'cell_row': 4, 'cell_col':3, 'cell_value':'Audit' }, {'cell_row': 4, 'cell_col':4, 'cell_value':'2019'} ]"

  - name: Update Employee Data
    register: result
    open_excel:
      src: "SampleBook.xlsx"
      updates_matrix: "{{ updates_matrix }}"
      sheet_name: "Employee"         
      op: "w"
      cell_style: "{{ cell_style }}"


          
  - name: Read Excel File
    hosts: localhost
    connection: local
    gather_facts: no

    tasks:
      - name: Show Employee Information
        register: result
        open_excel:
          src: "SampleBook.xlsx"
          op: "r"
          index_by_name: True
          
      - debug: var=result
'''



import openpyxl


#########################################################
####     READ CONTENT: read_xl_content
#########################################################
def read_xl_content(excel_file, index_by_name, read_range, sheet_name):
    
    retval = {}
    excelsheet = {}


    ### Validate & Initialize Parameters
    try:
        start_row = read_range['start_row']
    except Exception as e:
        start_row = 1

    try:
        start_col = read_range['start_col']
    except Exception as e:
        start_col = 1

    try:
        end_row = int(read_range['end_row']) + 1
    except Exception as e:
        end_row = 0

    try:
        end_col = int(read_range['end_col']) + 1
    except Exception as e:
        end_col = 0

    
    
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet_index = 0
        
        sheet_names_list = [sheet_name]
        if not sheet_name:
            sheet_names_list = wb.get_sheet_names()
        
        for asheet_name in sheet_names_list:
            ansible_sheet_index_name = 'sheet_index_' + str(sheet_index)
            excelsheet[ansible_sheet_index_name] = []
            current_sheet = wb.get_sheet_by_name(asheet_name)
            
            if end_row == 0:
                end_row = current_sheet.max_row + 1

            if end_col == 0:
                end_col = current_sheet.max_column + 1                                        

            dict_keys = []
            for col in range(start_col, end_col):
                if index_by_name:
                    dict_keys.append(str(current_sheet.cell(row=1, column=col).value))
                else:
                    dict_keys.append('col_'+str(col))
            for row in range (start_row, end_row):
                temp_dict = {}
                for col in range(start_col, end_col):
                    temp_dict[dict_keys[col-start_col]] = str(current_sheet.cell(row=row, column=col).value)
                excelsheet[ansible_sheet_index_name].append(temp_dict)
            sheet_index += 1
                        
    except IOError:
        return (1, "Error accessing excel file [%s]" % excel_file)

    retval = excelsheet
    
    return (0, retval)


#########################################################
###     WRITE CONTENT: update_xl_content
#########################################################
def update_xl_content(excel_file, dest_filename, updates_matrix, cell_style, sheet_name, op):
    
    try:
        if not dest_filename:
            dest_filename = excel_file + '_updated.xlsx'
            
        wb = openpyxl.load_workbook(excel_file, data_only=True)
    except IOError:
        return (1, "Error accessing excel file [%s]" % excel_file)

    try:
        current_sheet = wb.get_sheet_by_name(sheet_name)
        
        l_op = op
        row_no = int(updates_matrix[0]['cell_row'])
        
        if l_op == "w" and row_no == 0:
            l_op = "a"
        
        if l_op == "i":
            current_sheet.insert_rows(idx=row_no, amount=1)
            
        if l_op == "a":
            row_no = current_sheet.max_row + 1

        for cell in updates_matrix:
            if l_op == "w":
                row_no = int(cell['cell_row'])

            cellref = current_sheet.cell(row=row_no, column=int(cell['cell_col']))
            cellref.value = cell['cell_value']
            
            # styles values can be "True", "False", or undefined. If undefined, then leave the cell style intact
            if cell_style:
                if cell_style['fontColor']:
                    current_sheet.cell(row=row_no, column=int(cell['cell_col'])).font = openpyxl.styles.Font(color=cell_style['fontColor'])
                    
                font_strike = openpyxl.styles.Font()
                
                # try accessing the font strike values (may not have been specified by user)
                try:
                    if cell_style['bold'] == True:
                        font_strike.bold = True
                    elif cell_style['bold'] == False:
                        font_strike.bold = False
                except:
                    pass
                
                try:
                    if cell_style['italic'] == True:
                        font_strike.italic = True
                    elif cell_style['italic'] == False:
                        font_strike.italic = False
                except:
                    pass

                try:
                    if cell_style['underline'] == True:
                        font_strike.underline = 'single'
                    elif cell_style['underline'] == False:
                        font_strike.underline = 'none'
                except:
                    pass

                    
                try:
                    if font_strike:
                        current_sheet.cell(row=row_no, column=int(cell['cell_col'])).font = font_strike
                except:
                    pass

                if cell_style['bgColor']:
                    current_sheet.cell(row=row_no, column=int(cell['cell_col'])).fill = openpyxl.styles.PatternFill(fgColor=cell_style['bgColor'], fill_type = 'solid')
            
    except Exception as e:
        return (1, "Error accessing specified excel sheet [%s]: %s" % (sheet_name, str(e)) ) 
    
    try:
        wb.save(filename=dest_filename)
    except IOError:
        return (1, "Error updating excel file")
    
    nulldict = {}
    
    return (0, nulldict)

##################################################################################################################
##################################################################################################################
####     M A I N 
##################################################################################################################
##################################################################################################################

def main():
    module = AnsibleModule(argument_spec = dict(
             src = dict(required=True),
             dest = dict(required=False),
             op = dict(required=True),
             index_by_name = dict(type='bool',default=True,required=False),
             read_range = dict(type='dict',required=False),
             updates_matrix = dict(type='list',required=False),
             cell_style = dict(type='dict',required=False),
             sheet_name = dict(required=False)
             ),
             add_file_common_args=True)
    
    ret_code = 0
    op = module.params["op"]
    
    if op == "r":
        ret_code, response = read_xl_content(module.params["src"], module.params["index_by_name"], module.params["read_range"], module.params["sheet_name"])
    elif op == "w" or op == "a" or op == "i":
        try:
            dest_filename = module.params["dest"]
        except e:
            dest_filename = ""
        
        # sheet_name has to be specified for "write" operation
        if not module.params["sheet_name"]:
            ret_code = 1
            response = "sheet_name has to be specified for 'w', 'i', and 'a' operation"
            
        if not ret_code:
            ret_code, response = update_xl_content(module.params["src"], dest_filename, module.params["updates_matrix"],  module.params['cell_style'], module.params["sheet_name"], op)
    else:
        ret_code = 1
        response = "Invalid Excel file access mode. Valid modes are: 'r', 'w', 'i', and 'a'"
    
    if ret_code:
        module.fail_json(msg=response)
    else:
        module.exit_json(changed=True, **response)

    return ret_code


from ansible.module_utils.basic import *
main()
#
